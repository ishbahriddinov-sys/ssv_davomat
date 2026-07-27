"""Формирование отчётов: Excel, CSV, PDF."""
from __future__ import annotations

import csv
import io
from datetime import date

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core import security
from app.db.models import Attendance, Department, User
from app.services import attendance_service


async def _collect_rows(
    session: AsyncSession, start: date, end: date, department_id: int | None = None
) -> list[dict]:
    stmt = (
        select(Attendance, User, Department)
        .join(User, Attendance.user_id == User.id)
        .join(Department, User.department_id == Department.id, isouter=True)
        .where(Attendance.work_date >= start, Attendance.work_date <= end)
    )
    if department_id:
        stmt = stmt.where(User.department_id == department_id)
    stmt = stmt.order_by(Attendance.work_date, User.id)

    rows: list[dict] = []
    for att, user, dept in (await session.execute(stmt)).all():
        rows.append(
            {
                "date": att.work_date.isoformat(),
                "corporate_id": user.corporate_id,
                "full_name": security.decrypt(user.full_name_enc) or "",
                "department": dept.name if dept else "",
                "position": user.position or "",
                "check_in": att.check_in.strftime("%H:%M") if att.check_in else "",
                "late_minutes": att.late_minutes,
                "status": STATUS_UZ.get(att.status.value, att.status.value),
                "geo_verified": "ҳа" if att.geo_verified else "йўқ",
            }
        )
    return rows


_FONT_CANDIDATES = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/dejavu/DejaVuSans.ttf",
    "/Library/Fonts/Arial Unicode.ttf",
    "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
]
_cyrillic_font: str | None = None


def _register_cyrillic_font() -> str:
    """Регистрирует Unicode-шрифт для корректного вывода кириллицы в PDF.

    Возвращает имя шрифта; при отсутствии TTF — стандартный Helvetica.
    """
    global _cyrillic_font
    if _cyrillic_font:
        return _cyrillic_font
    import os

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    for path in _FONT_CANDIDATES:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("AttendanceUnicode", path))
                _cyrillic_font = "AttendanceUnicode"
                return _cyrillic_font
            except Exception:  # noqa: BLE001
                continue
    _cyrillic_font = "Helvetica"
    return _cyrillic_font


HEADERS = [
    "date", "corporate_id", "full_name", "department", "position",
    "check_in", "late_minutes", "status", "geo_verified",
]
# Заголовки в интерфейсе — на узбекском (кириллица)
HEADERS_UZ = [
    "Сана", "Корп. ID", "Ф.И.Ш.", "Бўлим", "Лавозим", "Келиш",
    "Кечикиш (дақ)", "Ҳолат", "Геолокация",
]

# Статусы посещаемости на узбекском (кириллица)
STATUS_UZ = {
    "present": "келди",
    "late": "кечикди",
    "absent": "келмади",
    "completed": "якунланди",
    "on_leave": "таътилда",
    "sick": "касал",
}


async def to_csv(
    session: AsyncSession, start: date, end: date, department_id: int | None = None
) -> bytes:
    rows = await _collect_rows(session, start, end, department_id)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(HEADERS_UZ)  # узбекская шапка (кириллица)
    for row in rows:
        writer.writerow([row[k] for k in HEADERS])
    return buf.getvalue().encode("utf-8-sig")


async def to_excel(
    session: AsyncSession, start: date, end: date, department_id: int | None = None
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = await _collect_rows(session, start, end, department_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    for col, title in enumerate(HEADERS_UZ, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows, start=2):
        for c, key in enumerate(HEADERS, start=1):
            ws.cell(row=r, column=c, value=row[key])

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def to_pdf(
    session: AsyncSession,
    start: date,
    end: date,
    department_id: int | None = None,
    title: str = "Давомат ҳисоботи",
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    font_name = _register_cyrillic_font()

    rows = await _collect_rows(session, start, end, department_id)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    for style_name in ("Title", "Normal"):
        styles[style_name].fontName = font_name
    elements = [
        Paragraph(f"{title}", styles["Title"]),
        Paragraph(f"Давр: {start} — {end}", styles["Normal"]),
        Spacer(1, 12),
    ]

    short_headers = ["Сана", "Ф.И.Ш.", "Бўлим", "Лавозим", "Келиш", "Кечикиш", "Ҳолат"]
    data = [short_headers]
    for row in rows:
        data.append([
            row["date"], row["full_name"][:24], row["department"][:18],
            row["position"][:18], row["check_in"],
            str(row["late_minutes"]), row["status"],
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF2FB")]),
        ])
    )
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()
